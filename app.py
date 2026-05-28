import streamlit as st
import pandas as pd
import openpyxl
import io
import re
import json
import os
from datetime import datetime, date
import zipfile

st.set_page_config(
    page_title="Facturación Dyaboo",
    page_icon="🛍️",
    layout="wide"
)

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def clean_nit(val):
    """Remove CC prefix, dots, spaces from NIT/cédula."""
    if val is None:
        return None
    s = str(val).strip()
    s = re.sub(r'^CC\s*', '', s, flags=re.IGNORECASE)
    s = s.replace('.', '').replace(',', '').replace(' ', '')
    # Remove scientific notation
    try:
        s = str(int(float(s)))
    except Exception:
        pass
    return s if s.isdigit() else s

def clean_phone(val):
    """Format phone as integer string, handle scientific notation."""
    if val is None or str(val).strip() == '':
        return ''
    try:
        return str(int(float(str(val))))
    except Exception:
        return str(val).strip()

def title_case_address(s):
    if not s:
        return ''
    return str(s).title()

def to_upper(s):
    if not s:
        return ''
    return str(s).upper().strip()

def parse_color_talla_from_name(product_name):
    """
    Extract color and talla from Shopify product name.
    Format: 'Product Name REF - Color / Talla'
    Returns (color_str, talla_str)
    """
    if not product_name:
        return '', ''
    # Split by ' / ' to get talla
    parts = str(product_name).split(' / ')
    talla = parts[-1].strip() if len(parts) > 1 else ''
    # Color: everything after last ' - '
    before_slash = parts[0] if parts else str(product_name)
    dash_parts = before_slash.split(' - ')
    color = dash_parts[-1].strip() if len(dash_parts) > 1 else ''
    return color, talla

def parse_ref_from_sku(sku):
    """
    Extract reference from SKU (remove last color+talla suffix).
    SKU format: REF + TALLA_CODE + COLOR_CODE
    e.g. DF10067M24 → ref=DF10067, talla_code=M (or XL etc.), color_code=24
    We extract ref by matching known talla codes from the end.
    """
    if not sku or str(sku).strip() in ('', '0', 'nan'):
        return '', '', ''
    sku = str(sku).strip()
    # Known talla codes (longest first to avoid partial matches)
    tallas = ['XXL', 'XL', 'XS', 'U', 'S', 'M', 'L',
              '6', '8', '10', '12', '14', '16', '18', '20', '22',
              '6X', '8X', '10X', '12X', '14X', '16X']
    for t in sorted(tallas, key=len, reverse=True):
        pattern = rf'^(.*?)({re.escape(t)})(\w+)$'
        m = re.match(pattern, sku)
        if m:
            ref = m.group(1)
            talla_code = m.group(2)
            color_code = m.group(3)
            return ref, talla_code, color_code
    # Fallback: last 2 chars color, rest ref
    return sku[:-2], '', sku[-2:]

def build_cod_articulo(ref):
    """Build SAG article code: ref + D, prefix 20 if starts with digit."""
    if not ref:
        return ''
    ref = str(ref).strip()
    if ref and ref[0].isdigit():
        return '20' + ref + 'D'
    return ref + 'D'

def get_vendedor(source, is_first_item):
    """202 for first item of web orders, 302 otherwise."""
    if is_first_item and str(source).lower() == 'web':
        return '202'
    return '302'

def get_flete_code(flete_val):
    """Return SAG flete article code based on flete value."""
    try:
        v = float(flete_val)
    except Exception:
        return None, None
    if v == 15000:
        return '15', round(15000 / 1.19, 8)
    elif v == 10000:
        return '10', round(10000 / 1.19, 8)
    elif v > 0:
        return 'OTRO', v
    return None, None

# ─────────────────────────────────────────────
# LOAD REFERENCE DATA FROM PLANTILLA
# ─────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def load_reference_data(plantilla_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(plantilla_bytes), read_only=True, data_only=True)

    # CIUDADES → {nombre_upper: zona}
    ws = wb['Ciudades']
    ciudades = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        ciudad_raw = row[6]  # 'Ciudad' column (index 6)
        zona = row[8]        # 'Zona' column (index 8)
        if ciudad_raw:
            ciudades[str(ciudad_raw).upper().strip()] = str(zona).strip() if zona else ''
    
    # TABLA DE COLORES → {color_name_upper: code}
    ws = wb['Tabla de Colores']
    color_map = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        name = row[0]
        code = row[1]
        if name:
            color_map[str(name).upper().strip()] = str(code).strip() if code else ''
    
    # COLORES SAG → {color_name_upper: code} (additional source)
    ws = wb['Colores']
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        name = row[1]   # SS_COLOR_LARGO
        code = row[2]   # SS_COLOR
        if name and str(name).upper().strip() not in color_map:
            color_map[str(name).upper().strip()] = str(code).strip() if code else ''
    
    # TALLAS → {talla_code: talla_code} (just validate they exist)
    ws = wb['Tallas']
    tallas_validas = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[1]:
            tallas_validas.add(str(row[1]).strip())
    
    # TERCEROS → set of NITs already in SAG
    ws = wb['Terceros']
    terceros = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[1]:
            terceros.add(str(int(float(str(row[1])))).strip())
    
    wb.close()
    return ciudades, color_map, tallas_validas, terceros


# ─────────────────────────────────────────────
# PROCESS SHOPIFY EXPORT
# ─────────────────────────────────────────────

def process_orders(df, ciudades, color_map, tallas_validas, terceros, fecha_facturacion):
    """
    Process Shopify orders dataframe and return:
    - orders_processed: list of dicts (one per line item)
    - errors: list of error dicts
    - new_clients: list of new client dicts
    """
    errors = []
    orders_processed = []
    new_clients = {}  # nit → client dict

    # Group by order Name
    order_groups = {}
    for _, row in df.iterrows():
        name = str(row.get('Name', '')).strip()
        if name not in order_groups:
            order_groups[name] = []
        order_groups[name].append(row)

    for order_name, items in order_groups.items():
        first_row = items[0]

        # ── NIT ──
        billing_company = first_row.get('Billing Company', '')
        nit_raw = str(billing_company).strip() if pd.notna(billing_company) else ''
        nit = clean_nit(nit_raw)

        # Skip "Envío Nacional" lines (SKU=0)
        items = [r for r in items if str(r.get('Lineitem sku', '')).strip() not in ('0', '', 'nan') 
                 and 'envío' not in str(r.get('Lineitem name', '')).lower()
                 and 'envio' not in str(r.get('Lineitem name', '')).lower()]
        if not items:
            continue

        # ── City / Zona ──
        billing_city = str(first_row.get('Billing City', '')).strip()
        # If city looks like a number (cédula stored in city field)
        city_key = billing_city.upper().strip()
        if city_key.isdigit():
            errors.append({
                'Pedido': order_name, 'NIT': nit,
                'Campo': 'Ciudad', 'Valor': billing_city,
                'Error': 'La ciudad parece una cédula. ¿Cuál es la ciudad real?',
                'Corrección': ''
            })
            ciudad_sag = ''
            zona_sag = ''
        else:
            # Try exact match, then partial
            found_city = ciudades.get(city_key, None)
            if found_city is None:
                # Try partial / with accent variants
                for k, v in ciudades.items():
                    if city_key in k or k in city_key:
                        found_city = v
                        city_key = k
                        break
            if found_city is None:
                errors.append({
                    'Pedido': order_name, 'NIT': nit,
                    'Campo': 'Ciudad', 'Valor': billing_city,
                    'Error': f'Ciudad "{billing_city}" no encontrada en tabla.',
                    'Corrección': ''
                })
                ciudad_sag = city_key
                zona_sag = ''
            else:
                ciudad_sag = city_key
                zona_sag = found_city

        # ── Client data ──
        billing_name = str(first_row.get('Billing Name', '')).strip()
        email = str(first_row.get('Email', '')).strip() if pd.notna(first_row.get('Email')) else '(en blanco)'
        phone = clean_phone(first_row.get('Billing Phone'))
        address = title_case_address(first_row.get('Billing Street', ''))
        source = str(first_row.get('Source', '')).strip()
        payment_method = str(first_row.get('Payment Method', '')).strip()
        flete_val = first_row.get('Shipping', 0)
        try:
            flete_val = float(str(flete_val).strip())
        except Exception:
            flete_val = 0

        # ── Is new client? ──
        if nit and nit not in terceros and nit not in new_clients:
            # Parse name
            parts = billing_name.split()
            ap1 = ap2 = n1 = n2 = ''
            if len(parts) >= 4:
                ap1, ap2, n1, n2 = parts[0], parts[1], parts[2], parts[3]
            elif len(parts) == 3:
                ap1, ap2, n1 = parts[0], parts[1], parts[2]
            elif len(parts) == 2:
                ap1, n1 = parts[0], parts[1]
            elif len(parts) == 1:
                n1 = parts[0]

            new_clients[nit] = {
                'NIT': nit,
                'Nombre Shopify': billing_name,
                'Nombre SAG (Validado)': '',
                'Ciudad': billing_city,
                'Departamento': str(first_row.get('Billing Province Name', '')).strip(),
                'Dirección': address,
                'Teléfono': phone,
                'Email': email if email else '(en blanco)',
                'Zona': zona_sag,
                'Ciudad SAG': ciudad_sag,
                'Apellido1': to_upper(ap1),
                'Apellido2': to_upper(ap2),
                'Nombre1': to_upper(n1),
                'Nombre2': to_upper(n2),
                'Observaciones': '',
                '_excluido': False
            }

        # ── Process line items ──
        flete_asignado = False
        for idx, item in enumerate(items):
            is_first = idx == 0
            sku = str(item.get('Lineitem sku', '')).strip()
            product_name = str(item.get('Lineitem name', '')).strip()

            # Parse color and talla
            color_str, talla_str = parse_color_talla_from_name(product_name)
            ref, talla_code_from_sku, color_code_from_sku = parse_ref_from_sku(sku)

            # Talla: prefer from product name, fallback to SKU
            talla_final = talla_str if talla_str else talla_code_from_sku

            # Color code lookup
            color_key = color_str.upper().strip()
            color_code = color_map.get(color_key, '')
            if not color_code:
                color_code = color_code_from_sku
                if not color_code:
                    errors.append({
                        'Pedido': order_name, 'NIT': nit,
                        'Campo': 'Color', 'Valor': color_str,
                        'Error': f'Color "{color_str}" no tiene código en tabla. SKU: {sku}',
                        'Corrección': ''
                    })

            cod_articulo = build_cod_articulo(ref)

            # Prices
            try:
                valor = float(str(item.get('Lineitem price', 0)).strip())
            except Exception:
                valor = 0
            try:
                base_valor = float(str(item.get('Lineitem compare at price', 0)).strip())
            except Exception:
                base_valor = valor

            # Without IVA
            valor_sin_iva = round(valor / 1.19, 8)
            base_sin_iva = round(base_valor / 1.19, 8) if base_valor else valor_sin_iva
            descuento = round(base_sin_iva - valor_sin_iva, 8) if base_sin_iva > valor_sin_iva else 0

            # Flete
            flete_linea = 0
            flete_code = None
            flete_valor_sin_iva = 0
            if is_first and flete_val > 0 and not flete_asignado:
                fc, fv = get_flete_code(flete_val)
                if fc == 'OTRO':
                    errors.append({
                        'Pedido': order_name, 'NIT': nit,
                        'Campo': 'Flete', 'Valor': flete_val,
                        'Error': f'Flete de ${flete_val:,.0f} no es $10.000 ni $15.000. ¿Qué código SAG usar?',
                        'Corrección': ''
                    })
                flete_code = fc
                flete_valor_sin_iva = fv or 0
                flete_linea = flete_val
                flete_asignado = True

            vendedor = get_vendedor(source, is_first)
            observacion = f"Pedido Dyaboo Online {order_name} Método de Pago: {payment_method}"

            orders_processed.append({
                '_order_name': order_name,
                '_nit': nit,
                '_excluir': False,
                'CÓD. FUENTE': 'PW',
                'NÚMERO': '',  # filled later with consecutivo
                'FECHA': fecha_facturacion.strftime('%d/%m/%Y'),
                'NIT': nit,
                'OBSERVACIONES': observacion,
                'COD.ARTICULO': cod_articulo,
                'COD. BODEGA': '09',
                'CANTIDAD': int(item.get('Lineitem quantity', 1)),
                'VALOR': valor_sin_iva,
                'IVA': 19,
                'DSCTO': round(descuento, 2) if descuento else 0,
                'NIT VENDEDOR': vendedor,
                'TALLA': talla_final,
                'COLOR': color_code,
                '_flete_val': flete_val,
                '_flete_code': flete_code,
                '_flete_valor_sin_iva': flete_valor_sin_iva,
                '_is_first': is_first,
                '_color_str': color_str,
                '_sku': sku,
                '_product_name': product_name,
            })

    return orders_processed, errors, new_clients


def assign_consecutivos(orders_processed, excluded_nits, start_num):
    """Assign consecutive SAG numbers, skipping excluded clients."""
    order_map = {}
    counter = start_num
    result = []
    for line in orders_processed:
        nit = line['_nit']
        order = line['_order_name']
        if nit in excluded_nits:
            continue
        if order not in order_map:
            order_map[order] = counter
            counter += 1
        line = dict(line)
        line['NÚMERO'] = order_map[order]
        result.append(line)
    return result, counter - 1


# ─────────────────────────────────────────────
# BUILD OUTPUT FILES
# ─────────────────────────────────────────────

def build_clientes_a_revisar(new_clients):
    rows = []
    for nit, c in new_clients.items():
        rows.append({
            'Cédula': nit,
            'Nombre Shopify': c['Nombre Shopify'],
            'Departamento': c['Departamento'],
            'Ciudad': c['Ciudad'],
            'Dirección': c['Dirección'],
            'Teléfono': c['Teléfono'],
            'Nombre SAG (Validado)': c['Nombre SAG (Validado)'],
            'Observaciones': c['Observaciones'],
        })
    return pd.DataFrame(rows)


def build_clientes_a_crear(new_clients, excluded_nits, fecha):
    rows = []
    for nit, c in new_clients.items():
        if nit in excluded_nits:
            continue
        nombre_sag = c['Nombre SAG (Validado)'] or to_upper(c['Nombre Shopify'])
        rows.append({
            'NIT ': nit,
            ' DV ': '',
            ' CÓDIGO ALTERNO ': '',
            ' NOMBRE ': nombre_sag,
            ' NOMBRE ALTERNO ': '',
            ' TELÉFONO PPAL ': c['Teléfono'],
            ' TELÉFONO ALTERNO ': '',
            ' FAX ': '',
            ' DIRECCIÓN ': c['Dirección'],
            ' CIUDAD ': c['Ciudad SAG'],
            ' NATURALEZA ': 'Natural',
            ' TIPO DE TERCERO ': '',
            'E_MAIL ': c['Email'],
            ' PÁGINA WEB ': '',
            'ACT. COMERCIAL ': '05-VENTA ONLINE',
            ' FORMA DE PAGO NORMAL ': 'CONTADO',
            ' CUPO ': '',
            ' ACTIVO (S/N) ': 'S',
            ' AUTORRETENEDOR (S/N) ': 'N',
            ' OBSERVACIONES ': '',
            ' TIPO DE RETEFTE NORMAL ': 'VO - NINGUNA -CXC-',
            ' TIPO DE CONTACTO ': '',
            ' NOMBRE DEL CONTACTO ': '',
            ' PRECIO DE VENTA (0..4) ': 4,
            ' NIT VENDEDOR ': 302,
            ' ZONA ': c['Zona'],
            ' CLASE DE CLIENTE ': 'V',
            ' CODIGO INTERNO ': '',
            ' Tipo de Dcto: Nit(A)/Cédula(C)/Cédula de Extr. (E)/ Cédula del Tutor(T)/Número Único de Identificación Personal NUIP(U) ': 'C',
            ' IVA (S/N) ': 'S',
            ' Dia Cumpleaños(01, 02, 03,04, ...)': '',
            ' Mes Nacimiento (01,02,03..12) ': '',
            ' CATEGORIA ': '',
            ' DSCTO COMERCIAL ': '',
            ' DSCTO PP ': '',
            ' NOMBRE1 ': c['Nombre1'],
            ' NOMBRE2 ': c['Nombre2'],
            ' APELLIDO1 ': c['Apellido1'],
            ' APELLIDO2 ': c['Apellido2'],
            ' DIRECCION2 ': '',
            ' RUT  ': '',
        })
    return pd.DataFrame(rows)


def build_clientes_complemento(new_clients, excluded_nits, fecha):
    rows = []
    for nit, c in new_clients.items():
        if nit in excluded_nits:
            continue
        nombre_sag = c['Nombre SAG (Validado)'] or to_upper(c['Nombre Shopify'])
        rows.append({
            'Nit ': nit,
            ' Digito Verificacion ': '',
            ' Tipo de Documento(C:Cedula, A:Nit) ': 'C',
            ' Nombre ': nombre_sag,
            ' Nombre1 ': c['Nombre1'],
            ' Nombre2 ': c['Nombre2'],
            ' Apellido1 ': c['Apellido1'],
            ' Apellido2 ': c['Apellido2'],
            ' Dirección ': c['Dirección'],
            ' Ciudad ': c['Ciudad SAG'],
            ' Telefono ppal ': c['Teléfono'],
            '     Código postal ': '000000',
            ' Email1': c['Email'],
            ' Email2(FE) ': c['Email'],
            ' Genera FE(S/N) ': 'S',
            ' Quien Aprueba ': 'LA WEB',
            ' Fecha Aprueba ': fecha.strftime('%d/%m/%Y'),
            ' Responsabilidad fiscal': 'R-99-PN',
            ' Nit Vendedor ': 302,
            'Clase Regimen (N:persona natural, G:Gran Contribuyente, R:No Responsable de IVA, C:Responsable de IVA Régimen Común, F:Régimen SIMPLE)': 'N',
        })
    return pd.DataFrame(rows)


def build_pedidos_df(orders_final):
    rows = []
    for line in orders_final:
        rows.append({
            'CÓD. FUENTE ': line['CÓD. FUENTE'],
            ' NÚMERO ': line['NÚMERO'],
            ' FECHA ': line['FECHA'],
            ' NIT ': line['NIT'],
            ' OBSERVACIONES ': line['OBSERVACIONES'],
            ' COD.ARTICULO ': line['COD.ARTICULO'],
            ' COD. BODEGA ': line['COD. BODEGA'],
            ' CANTIDAD ': line['CANTIDAD'],
            ' VALOR ': line['VALOR'],
            ' IVA ': line['IVA'],
            ' DSCTO ': line['DSCTO'],
            ' NIT VENDEDOR ': line['NIT VENDEDOR'],
            'TALLA': line['TALLA'],
            'COLOR': line['COLOR'],
        })
        # Flete line
        if line.get('_flete_code') and line['_flete_code'] not in (None, 'OTRO'):
            rows.append({
                'CÓD. FUENTE ': line['CÓD. FUENTE'],
                ' NÚMERO ': line['NÚMERO'],
                ' FECHA ': line['FECHA'],
                ' NIT ': line['NIT'],
                ' OBSERVACIONES ': line['OBSERVACIONES'],
                ' COD.ARTICULO ': line['_flete_code'],
                ' COD. BODEGA ': line['COD. BODEGA'],
                ' CANTIDAD ': 1,
                ' VALOR ': line['_flete_valor_sin_iva'],
                ' IVA ': 19,
                ' DSCTO ': 0,
                ' NIT VENDEDOR ': line['NIT VENDEDOR'],
                'TALLA': 'SURT',
                'COLOR': 'SURT',
            })
    return pd.DataFrame(rows)


def df_to_excel_bytes(df, sheet_name='Sheet1'):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return buf.getvalue()


def df_to_csv_bytes(df):
    return df.to_csv(index=False, sep=';', encoding='utf-8-sig').encode('utf-8-sig')


# ─────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────

st.title("🛍️ Facturación Dyaboo — SAG")
st.caption("Convierte pedidos de Shopify al formato de carga SAG")

# ── Sidebar: Plantilla (persistent reference) ──
with st.sidebar:
    st.header("⚙️ Configuración")
    st.subheader("Plantilla SAG")
    st.caption("Solo cárgala cuando actualices Terceros, Ciudades o Colores.")

    plantilla_key = "plantilla_bytes"
    plantilla_date_key = "plantilla_date"

    uploaded_plantilla = st.file_uploader(
        "Subir Plantilla.xlsx", type=['xlsx'], key='plantilla_uploader'
    )
    if uploaded_plantilla:
        st.session_state[plantilla_key] = uploaded_plantilla.read()
        st.session_state[plantilla_date_key] = datetime.now().strftime('%d/%m/%Y %H:%M')
        st.success("Plantilla guardada.")

    if plantilla_key in st.session_state:
        st.info(f"📋 Plantilla activa desde: {st.session_state.get(plantilla_date_key, '—')}")
    else:
        st.warning("Sin plantilla. Carga una para continuar.")

    st.divider()
    st.subheader("Parámetros del lote")
    consecutivo_inicio = st.number_input(
        "Consecutivo inicial (último lote + 1)",
        min_value=1, value=5388, step=1
    )
    fecha_facturacion = st.date_input(
        "Fecha de facturación",
        value=date.today()
    )

# ── Main: Upload Shopify export ──
if plantilla_key not in st.session_state:
    st.info("👈 Primero carga la Plantilla SAG en el panel izquierdo.")
    st.stop()

# Load reference data
try:
    with st.spinner("Cargando tablas de referencia..."):
        ciudades, color_map, tallas_validas, terceros = load_reference_data(
            st.session_state[plantilla_key]
        )
except Exception as e:
    st.error(f"Error leyendo la plantilla: {e}")
    st.stop()

st.subheader("📂 Paso 1 — Cargar pedidos de Shopify")
uploaded_orders = st.file_uploader(
    "Subir export de Shopify (.xlsx o .csv)",
    type=['xlsx', 'csv'],
    key='orders_uploader'
)

if not uploaded_orders:
    st.info("Sube el archivo de pedidos exportado desde Shopify para continuar.")
    st.stop()

# Read orders
try:
    if uploaded_orders.name.endswith('.csv'):
        df = pd.read_csv(uploaded_orders)
    else:
        df = pd.read_excel(uploaded_orders)
    df = df.where(pd.notna(df), None)
except Exception as e:
    st.error(f"Error leyendo el archivo de pedidos: {e}")
    st.stop()

st.success(f"✅ {len(df)} filas cargadas — {df['Name'].nunique() if 'Name' in df.columns else '?'} pedidos")

# Process
with st.spinner("Analizando pedidos..."):
    orders_processed, errors, new_clients = process_orders(
        df, ciudades, color_map, tallas_validas, terceros, fecha_facturacion
    )

# ── Step 2: Show errors ──
st.subheader("🔍 Paso 2 — Revisión de errores")

if errors:
    st.warning(f"Se encontraron **{len(errors)} errores** que necesitan corrección:")
    err_df = pd.DataFrame(errors)
    
    # Editable corrections
    st.caption("Llena la columna **Corrección** para los errores que puedas resolver:")
    edited = st.data_editor(err_df, use_container_width=True, num_rows="fixed",
                             column_config={"Corrección": st.column_config.TextColumn(width="large")})
    
    # Apply corrections back
    corrections = {}
    for _, row in edited.iterrows():
        if row['Corrección']:
            corrections[(row['Pedido'], row['Campo'])] = row['Corrección']
    
    # Apply city corrections
    for (pedido, campo), correccion in corrections.items():
        if campo == 'Ciudad':
            city_key = correccion.upper().strip()
            zona = ciudades.get(city_key, '')
            for line in orders_processed:
                if line['_order_name'] == pedido:
                    pass  # City correction would need re-processing; flag for user
    
    st.info("💡 Los pedidos con errores no corregidos de **cédula incorrecta** deben marcarse como 'INCORRECTO' en el paso de clientes a revisar.")
else:
    st.success("✅ Sin errores detectados. Puedes continuar.")

# ── Step 3: New clients ──
st.subheader("👤 Paso 3 — Clientes nuevos")

if not new_clients:
    st.success("✅ Todos los clientes ya existen en el SAG. No hay clientes a crear.")
    excluded_nits = set()
else:
    st.info(f"Se encontraron **{len(new_clients)} clientes nuevos** para crear en el SAG.")

    # Show editable table for police review
    st.markdown("**Clientes a Revisar** — descarga, valida en policía y marca los incorrectos:")
    revisar_df = build_clientes_a_revisar(new_clients)
    
    st.download_button(
        "⬇️ Descargar Clientes a Revisar",
        data=df_to_excel_bytes(revisar_df),
        file_name=f"Clientes_a_Revisar_{fecha_facturacion.strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.markdown("**Sube el archivo revisado** (con nombres validados y observaciones):")
    uploaded_revisado = st.file_uploader(
        "Subir Clientes_a_Revisar revisado", type=['xlsx'], key='revisado_uploader'
    )

    excluded_nits = set()
    if uploaded_revisado:
        rev_df = pd.read_excel(uploaded_revisado)
        for _, row in rev_df.iterrows():
            nit = str(row.get('Cédula', '')).strip()
            obs = str(row.get('Observaciones', '')).strip().upper()
            nombre_validado = str(row.get('Nombre SAG (Validado)', '')).strip()
            if nit in new_clients:
                if 'INCORRECTO' in obs:
                    excluded_nits.add(nit)
                    new_clients[nit]['_excluido'] = True
                else:
                    if nombre_validado:
                        new_clients[nit]['Nombre SAG (Validado)'] = nombre_validado
                    new_clients[nit]['Observaciones'] = obs

        valid_count = len(new_clients) - len(excluded_nits)
        if excluded_nits:
            st.warning(f"⚠️ {len(excluded_nits)} cliente(s) marcados como INCORRECTO — sus pedidos serán excluidos.")
        st.success(f"✅ {valid_count} clientes válidos para crear.")
    else:
        st.info("Sube el archivo revisado para continuar, o continúa sin revisión (todos los clientes serán incluidos).")

# ── Step 4: Generate files ──
st.subheader("📦 Paso 4 — Generar archivos SAG")

orders_final, ultimo_consecutivo = assign_consecutivos(
    orders_processed, excluded_nits, consecutivo_inicio
)

col1, col2, col3 = st.columns(3)
with col1:
    st.metric("Pedidos procesados", len(set(l['_order_name'] for l in orders_processed)))
with col2:
    st.metric("Pedidos excluidos", len(excluded_nits))
with col3:
    st.metric(f"Consecutivos usados", f"{consecutivo_inicio} → {ultimo_consecutivo}")

if not orders_final:
    st.warning("No hay pedidos para generar.")
    st.stop()

pedidos_df = build_pedidos_df(orders_final)

st.markdown("**Vista previa del plano de pedidos:**")
st.dataframe(pedidos_df.head(20), use_container_width=True)

# Download section
st.markdown("### ⬇️ Descargar archivos")

fecha_str = fecha_facturacion.strftime('%Y%m%d')

files_to_zip = {}

# Pedidos
pedidos_xlsx = df_to_excel_bytes(pedidos_df, 'PlanoPedidoSAG_TallaColor')
pedidos_csv = df_to_csv_bytes(pedidos_df)
files_to_zip[f'Pedidos_a_Cargar_{fecha_str}.xlsx'] = pedidos_xlsx
files_to_zip[f'Pedidos_a_Cargar_{fecha_str}.csv'] = pedidos_csv

col1, col2 = st.columns(2)
with col1:
    st.download_button("⬇️ Pedidos a Cargar (XLSX)", pedidos_xlsx,
                       f"Pedidos_a_Cargar_{fecha_str}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
with col2:
    st.download_button("⬇️ Pedidos a Cargar (CSV)", pedidos_csv,
                       f"Pedidos_a_Cargar_{fecha_str}.csv", mime="text/csv")

# Clients (only if there are new ones validated)
valid_clients = {k: v for k, v in new_clients.items() if k not in excluded_nits}
if valid_clients and uploaded_revisado:
    crear_df = build_clientes_a_crear(new_clients, excluded_nits, fecha_facturacion)
    complemento_df = build_clientes_complemento(new_clients, excluded_nits, fecha_facturacion)

    crear_xlsx = df_to_excel_bytes(crear_df, 'Cliente02')
    complemento_xlsx = df_to_excel_bytes(complemento_df, 'Clientes04-Complemento')
    files_to_zip[f'Clientes_a_Crear_{fecha_str}.xlsx'] = crear_xlsx
    files_to_zip[f'Clientes_Complementario_{fecha_str}.xlsx'] = complemento_xlsx

    col1, col2 = st.columns(2)
    with col1:
        st.download_button("⬇️ Clientes a Crear", crear_xlsx,
                           f"Clientes_a_Crear_{fecha_str}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with col2:
        st.download_button("⬇️ Clientes Complementario", complemento_xlsx,
                           f"Clientes_Complementario_{fecha_str}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ZIP with everything
zip_buf = io.BytesIO()
with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
    for fname, fbytes in files_to_zip.items():
        zf.writestr(fname, fbytes)

st.divider()
st.download_button(
    "📦 Descargar TODO en ZIP",
    data=zip_buf.getvalue(),
    file_name=f"Dyaboo_SAG_{fecha_str}.zip",
    mime="application/zip",
    type="primary",
    use_container_width=True
)

st.caption(f"Último consecutivo de este lote: **{ultimo_consecutivo}** — el próximo lote arranca en **{ultimo_consecutivo + 1}**")
